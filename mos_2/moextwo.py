import win32com.client
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl.styles import Alignment
import os
import requests


def curr_pars(url=None, moment_start=None, moment_end=None):
    dict_curr = {'USD_RUB': {}, 'EUR_RUB': {}}
    currency = ['USD_RUB', 'EUR_RUB']
    for cur in currency:
        params = {'language': 'en', 'currency': cur, 'moment_start': moment_start,
                  'moment_end': moment_end}
        response = requests.get(url, params=params)
        soup = BeautifulSoup(response.content, "xml")
        res = soup.find_all('rate')
        formater = '%Y-%m-%d %H:%M:%S'
        # Запись курса валют в словарь
        for i in res:
            curr_date = datetime.date(datetime.strptime(i['moment'], formater))
            # Если дата в словаре уже существует
            if curr_date in dict_curr[cur]:
                a = dict_curr[cur][curr_date][0]
                # округление курса до двух знаков
                b = round(float(i['value']), 2)
                # округление разницы до двух знаков
                curr_value = round((a - b), 2)
                dict_curr[cur][curr_date].append(curr_value)
            # Запись даты в словарь с последним курсом
            else:
                dict_curr[cur][curr_date] = [round(float(i['value']), 2)]
    return dict_curr


class Myclass:
    def __init__(self, dict_curr=None, address=None):
        self.file = os.path.abspath(r'Динамика.xlsx')
        self.address = address
        self.dict_curr = dict_curr

    def rec_excl(self):
        wb = openpyxl.load_workbook(self.file)
        sheet = wb['Лист1']
        end_row = sheet.max_row

        for d in self.dict_curr:
            row_count = 2
            if d == 'EUR_RUB':
                count_colm = 3
                cell_date = 'D'
                cell_rate = 'E'
                cell_chang = 'F'
            else:
                count_colm = 0
                cell_date = 'A'
                cell_rate = 'B'
                cell_chang = 'C'

            # Запись значений в ячейки + форматирование
            for key in self.dict_curr[d]:
                sheet.cell(row=row_count, column=1 + count_colm).value = key
                # Форматирование
                sheet.cell(row=row_count, column=1 + count_colm).alignment = Alignment(horizontal='center')
                sheet.column_dimensions[cell_date].width = 14

                sheet.cell(row=row_count, column=2 + count_colm).value = self.dict_curr[d][key][0]
                # Форматирование
                sheet.cell(row=row_count, column=2 + count_colm).number_format = '#,##0.00₽'
                sheet.cell(row=row_count, column=2 + count_colm).alignment = Alignment(horizontal='center')
                sheet.column_dimensions[cell_rate].width = 14

                sheet.cell(row=row_count, column=3 + count_colm).value = self.dict_curr[d][key][1]
                # Форматирование
                sheet.cell(row=row_count, column=3 + count_colm).number_format = '#,##0.00₽'
                sheet.cell(row=row_count, column=3 + count_colm).alignment = Alignment(horizontal='center')
                sheet.column_dimensions[cell_chang].width = 14
                row_count += 1

        # Расчет отношения EUR/USD в столбце G, если все столбцы заполнены в таблице
        if sheet['B2'].value is not None and sheet['E2'].value is not None:
            for row in range(2, end_row + 1):
                rate_USD = sheet.cell(row=row, column=2).value
                rate_EUR = sheet.cell(row=row, column=5).value
                sheet.cell(row=row, column=7).value = round(rate_EUR / rate_USD, 4)
                # Форматирование
                sheet.cell(row=row, column=7).alignment = Alignment(horizontal='center')
                sheet.column_dimensions['G'].width = 14
            wb.save(os.path.abspath(self.file))
            return end_row - 1

    def sander(self):
        path = os.path.abspath(self.file)
        wb = openpyxl.load_workbook(os.path.abspath(self.file))
        sheet = wb['Лист1']
        count_row = sheet.max_row - 1
        message = 'строк' + (
            '' if count_row % 10 == 0 or 4 < count_row % 10 < 10 or 10 < count_row % 100 < 15 else 'и' if 1 < count_row % 10 < 5 else 'a')
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = self.address
        mail.Subject = 'Динамика курса валют'
        mail.Body = 'Добавлено {} {}'.format(count_row, message)
        attachment = path
        mail.Attachments.Add(attachment)
        mail.Send()
        print('Соообщение отправлено {}'.format(self.address))


if __name__ == '__main__':
    url = 'https://moex.com/export/derivatives/currency-rate.aspx?'
    moment_start = '2021-07-09'
    moment_end = '2021-08-09'

    dict_curr = curr_pars(url=url, moment_start=moment_start, moment_end=moment_end)
    address = 'tcyganov_sa@nniirt.ru'
    obj = Myclass(dict_curr, address)
    obj.rec_excl()
    # obj.sander()
