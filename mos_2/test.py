import unittest
from moextwo import curr_pars, Myclass
import openpyxl
import os

class TestMos(unittest.TestCase):
    url = 'https://moex.com/export/derivatives/currency-rate.aspx?'
    moment_start = '2021-07-09'
    moment_end = '2021-08-09'
    dict_curr = curr_pars(url, moment_start, moment_end)
    wb = openpyxl.load_workbook(os.path.abspath(r'Динамика.xlsx'))
    sheet = wb['Лист1']
    end_row = sheet.max_row
    len_USD = len(dict_curr['USD_RUB'])
    len_EUR = len(dict_curr['EUR_RUB'])

    # Проверка равенства записей курсов в запросе и кол-ва записей в файле Excel
    # Проверка равенства записей курса USD и EUR полученных в запросе request
    def test_len(self):
        self.assertNotEqual(self.dict_curr['USD_RUB'], {})
        self.assertNotEqual(self.dict_curr['EUR_RUB'], {})
        self.assertEqual(len(self.dict_curr['USD_RUB']), len(self.dict_curr['EUR_RUB']))
        self.assertEqual(Myclass(dict_curr=self.dict_curr).rec_excl(), len(self.dict_curr['USD_RUB']))



    # Проверка правильности адресации записей в файле excel
    def test_excl(self):
        for i in range(1, 8):
            for j in range(1, 8):
                if i == j:
                    continue
                else:
                    self.assertIsNot(self.sheet.cell(row=self.end_row, column=i).value,
                                     self.sheet.cell(row=self.end_row, column=j).value)

        for i in range(1, 8):
            for j in range(1, 8):
                if i == j:
                    continue
                else:
                    self.assertIsNone(self.sheet.cell(row=self.end_row + 1, column=i).value,
                                      self.sheet.cell(row=self.end_row + 1, column=j).value)
