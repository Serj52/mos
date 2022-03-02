import win32com.client
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl.styles import Alignment
import os
import requests


def sander(file_name=None, adress=None):
    """
    Send function
    """
    path = os.path.abspath(file_name)
    wb = openpyxl.load_workbook(path)
    sheet = wb['Лист1']
    count_row = sheet.max_row - 1
    message = 'строк' + ('' if count_row % 10 == 0 or 4 < count_row % 10 < 10 or 10 < count_row % 100 < 15 else 'и' if 1 < count_row % 10 < 5 else 'a')
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = adress
    mail.Subject = 'Динамика курса валют'
    mail.Body = 'Добавлено {} {}'.format(count_row, message)
    attachment = path
    mail.Attachments.Add(attachment)
    mail.Send()


def rec_excl(dict_curr, currency):
    """Record function in Excel"""

    path = os.path.abspath('Динамика.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb['Лист1']
    end_row = sheet.max_row
    row_count = 2
    count_colm = 0
    cell_date = 'A'
    cell_rate = 'B'
    cell_chang = 'C'
    if currency == 'EUR_RUB':
        count_colm = 3
        cell_date = 'D'
        cell_rate = 'E'
        cell_chang = 'F'

    # Record in Excel and document formatting
    for key in dict_curr:
        sheet.cell(row=row_count, column=1 + count_colm).value = key

        sheet.cell(row=row_count, column=1 + count_colm).alignment = Alignment(horizontal='center')
        sheet.column_dimensions[cell_date].width = 14

        sheet.cell(row=row_count, column=2 + count_colm).value = dict_curr[key][0]

        sheet.cell(row=row_count, column=2 + count_colm).number_format = '#,##0.00₽'
        sheet.cell(row=row_count, column=2 + count_colm).alignment = Alignment(horizontal='center')
        sheet.column_dimensions[cell_rate].width = 14

        sheet.cell(row=row_count, column=3 + count_colm).value = dict_curr[key][1]

        sheet.cell(row=row_count, column=3 + count_colm).number_format = '#,##0.00₽'
        sheet.cell(row=row_count, column=3 + count_colm).alignment = Alignment(horizontal='center')
        sheet.column_dimensions[cell_chang].width = 14
        row_count += 1

    wb.save(path)
    # Calculate EUR/USD in column G, if all columns are field
    if sheet['B2'].value is not None and sheet['E2'].value is not None:
        for row in range(2, end_row+1):
            rate_USD = sheet.cell(row=row, column=2).value
            rate_EUR = sheet.cell(row=row, column=5).value
            sheet.cell(row=row, column=7).value = round(rate_EUR / rate_USD, 4)
            # Formatting
            sheet.cell(row=row, column=7).alignment = Alignment(horizontal='center')
            sheet.column_dimensions['G'].width = 14

        wb.save(path)
        count_row = end_row - 1
        return count_row

    else:
        count_row = end_row - 1
        return count_row

def curr_pars(path=None, params=None):
    """Get the exchange rate for sorted data"""

    response = requests.get(path, params=params)
    soup = BeautifulSoup(response.content, "xml")
    res = soup.find_all('rate')
    formater = '%Y-%m-%d %H:%M:%S'
    dict_curr = {}
    # Record exchange rate in dictionary
    for i in res:
        curr_date = datetime.date(datetime.strptime(i['moment'], formater))
        # If date in dictionary
        if curr_date in dict_curr:
            a = dict_curr[curr_date][0]
            # Rounding to two decimal places
            b = round(float(i['value']), 2)
            curr_value = round((a - b), 2)
            dict_curr[curr_date].append(curr_value)
        else:
            dict_curr[curr_date] = [round(float(i['value']), 2)]
    return dict_curr


def curr_pars_ver2(path=None, params=None):
    """Get the exchange rate for unsorted data"""

    response = requests.get(path, params=params)
    soup = BeautifulSoup(response.content, "xml")
    res = soup.find_all('rate')
    formater = '%Y-%m-%d %H:%M:%S'
    dict_curr = {}
    count = 0
    for i in res:
        curr_date = datetime.date(datetime.strptime(i['moment'], formater))
        if curr_date in dict_curr:
            a = dict_curr[curr_date][0]
            b = round(float(i['value']), 2)
            curr_value = round((a - b), 2)
            dict_curr[curr_date].append(curr_value)
            date_old = res[count]['moment']
            date_new = i['moment']
            # If current date later
            if date_new > date_old:
                dict_curr[curr_date] = [round(float(i['value']), 2)]
                dict_curr[curr_date].append(curr_value)
                count += 2
            else:
                count += 2
        else:
            dict_curr[curr_date] = [round(float(i['value']), 2)]
    return dict_curr



if __name__ == '__main__':
    currency = ['USD_RUB', 'EUR_RUB']
    for cur in currency:
        path = 'https://moex.com/export/derivatives/currency-rate.aspx?'
        params = {'language': 'en', 'currency': cur, 'moment_start': '2021-07-09',
                  'moment_end': '2021-08-09'}
        dict_curr = curr_pars(path=path, params=params)
        rec_excl(dict_curr=dict_curr, currency=cur)

    # Отправка файла по Outlook
    # sander(file_name='Динамика.xlsx',
    #        adress='tcyganov_sa@nniirt.ru')

